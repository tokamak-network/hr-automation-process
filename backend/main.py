import os
import json
import asyncio
from datetime import datetime
from contextlib import asynccontextmanager
from typing import Optional

from fastapi import FastAPI, HTTPException, Request, BackgroundTasks, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv

load_dotenv()

import sqlite3
from database import init_db, get_db, DB_PATH
from analyzer import analyze_repo, ai_analyze, analyze_github_profile, TEAM_MEMBERS, recommend_reviewers, calculate_weighted_score
from team_profiler import scan_org_profiles
from linkedin_google import search_linkedin_candidates, get_linkedin_candidates, update_candidate_status as update_linkedin_status, init_linkedin_db
from github_linkedin import bridge_github_candidates
from github_sourcing import search_github_developers
from matching import match_candidate_to_team

@asynccontextmanager
async def lifespan(app: FastAPI):
    await init_db()
    init_linkedin_db()
    yield

app = FastAPI(title="Tokamak Hiring Framework", lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


@app.get("/api/health")
async def health_check():
    return {"status": "healthy"}


def get_user_email(request: Request) -> Optional[str]:
    return request.headers.get("X-User-Email")


class CandidateSubmission(BaseModel):
    name: str
    email: str
    repo_url: str
    description: str = ""


# ---- User endpoints ----

@app.get("/api/users")
async def list_users():
    db = await get_db()
    rows = await db.execute("SELECT id, name, email, role FROM users ORDER BY name")
    users = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return users


@app.get("/api/users/me")
async def get_current_user(request: Request):
    email = get_user_email(request)
    if not email:
        return {"user": None}
    db = await get_db()
    row = await db.execute("SELECT id, name, email, role FROM users WHERE email = ?", (email,))
    user = await row.fetchone()
    await db.close()
    if not user:
        return {"user": None}
    return {"user": dict(user)}


# ---- Candidate endpoints ----

@app.post("/api/candidates/submit")
async def submit_candidate(data: CandidateSubmission):
    db = await get_db()
    cursor = await db.execute(
        "INSERT INTO candidates (name, email, repo_url, description) VALUES (?, ?, ?, ?)",
        (data.name, data.email, data.repo_url, data.description)
    )
    await db.commit()
    cid = cursor.lastrowid
    await db.close()
    return {"id": cid, "status": "submitted"}


@app.post("/api/candidates/{candidate_id}/analyze")
async def analyze_candidate(candidate_id: int, request: Request):
    user_email = get_user_email(request)
    db = await get_db()
    row = await db.execute("SELECT * FROM candidates WHERE id = ?", (candidate_id,))
    candidate = await row.fetchone()
    if not candidate:
        await db.close()
        raise HTTPException(404, "Candidate not found")

    repo_analysis = await analyze_repo(candidate["repo_url"])
    if "error" in repo_analysis:
        await db.close()
        raise HTTPException(400, repo_analysis["error"])

    ai_result = await ai_analyze(repo_analysis, candidate["description"])

    track_b = ai_result.get("track_b", {})
    weighted_score = ai_result.get("weighted_score", 0)

    await db.execute(
        """UPDATE candidates SET status='analyzed', scores=?, report=?, recommendation=?,
           repo_analysis=?, track_b_evaluation=?, weighted_score=?, analyzed_by=?, analyzed_at=? WHERE id=?""",
        (
            json.dumps(ai_result.get("scores", {})),
            ai_result.get("report", ""),
            ai_result.get("recommendation", "Maybe"),
            json.dumps({k: v for k, v in repo_analysis.items() if k != "sample_code"}),
            json.dumps(track_b),
            weighted_score,
            user_email or "",
            datetime.utcnow().isoformat(),
            candidate_id
        )
    )
    await db.commit()
    await db.close()
    return {
        "id": candidate_id,
        "status": "analyzed",
        "scores": ai_result.get("scores"),
        "weighted_score": weighted_score,
        "track_b": track_b,
        "recommendation": ai_result.get("recommendation"),
    }


@app.get("/api/candidates")
async def list_candidates():
    db = await get_db()
    rows = await db.execute(
        "SELECT id, name, email, repo_url, status, scores, recommendation, weighted_score, reviewed_by, analyzed_by, created_at FROM candidates ORDER BY created_at DESC"
    )
    candidates = []
    for r in await rows.fetchall():
        c = dict(r)
        if c["scores"]:
            c["scores"] = json.loads(c["scores"])
        candidates.append(c)
    await db.close()
    return candidates


@app.get("/api/candidates/{candidate_id}")
async def get_candidate(candidate_id: int):
    db = await get_db()
    row = await db.execute("SELECT * FROM candidates WHERE id = ?", (candidate_id,))
    candidate = await row.fetchone()
    await db.close()
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    c = dict(candidate)
    for field in ["scores", "repo_analysis", "track_b_evaluation"]:
        if c.get(field):
            c[field] = json.loads(c[field])
    return c


@app.get("/api/candidates/{candidate_id}/report")
async def get_report(candidate_id: int):
    db = await get_db()
    row = await db.execute("SELECT report, scores, recommendation, name, weighted_score, track_b_evaluation FROM candidates WHERE id = ?", (candidate_id,))
    candidate = await row.fetchone()
    await db.close()
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    c = dict(candidate)
    if c.get("scores"):
        c["scores"] = json.loads(c["scores"])
    if c.get("track_b_evaluation"):
        c["track_b_evaluation"] = json.loads(c["track_b_evaluation"])
    return c


@app.get("/api/candidates/{candidate_id}/recommended-reviewers")
async def get_recommended_reviewers(candidate_id: int, request: Request):
    user_email = get_user_email(request)
    db = await get_db()
    row = await db.execute("SELECT scores, repo_analysis FROM candidates WHERE id = ?", (candidate_id,))
    candidate = await row.fetchone()
    if not candidate:
        await db.close()
        raise HTTPException(404, "Candidate not found")

    scores = json.loads(candidate["scores"]) if candidate["scores"] else {}
    repo_analysis = json.loads(candidate["repo_analysis"]) if candidate["repo_analysis"] else {}

    reviewers = await recommend_reviewers(scores, repo_analysis, db, exclude_email=user_email)
    await db.close()
    return {"reviewers": reviewers}


@app.post("/api/candidates/{candidate_id}/review")
async def mark_reviewed(candidate_id: int, request: Request):
    user_email = get_user_email(request)
    if not user_email:
        raise HTTPException(400, "X-User-Email header required")
    db = await get_db()
    row = await db.execute("SELECT id FROM candidates WHERE id = ?", (candidate_id,))
    if not await row.fetchone():
        await db.close()
        raise HTTPException(404, "Candidate not found")
    await db.execute("UPDATE candidates SET reviewed_by = ? WHERE id = ?", (user_email, candidate_id))
    await db.commit()
    await db.close()
    return {"status": "reviewed", "reviewed_by": user_email}


# ---- Monitor endpoints ----

# Background scan state
_scan_status = {"running": False, "last_result": None, "last_error": None, "started_at": None}

@app.get("/api/monitor/scan/status")
async def scan_status():
    return _scan_status

@app.post("/api/monitor/scan")
async def scan_github(background_tasks: BackgroundTasks):
    token = os.getenv("GITHUB_TOKEN", "")
    if not token:
        raise HTTPException(400, "GITHUB_TOKEN not configured")
    if _scan_status["running"]:
        return {"status": "already_running", "started_at": _scan_status["started_at"]}
    _scan_status["running"] = True
    _scan_status["started_at"] = datetime.utcnow().isoformat()
    _scan_status["last_error"] = None
    background_tasks.add_task(_do_scan_github, token)
    return {"status": "started", "message": "Scan started in background. Check /api/monitor/scan/status for progress."}

async def _do_scan_github(token: str):
    try:
        result = await asyncio.to_thread(_scan_github_sync, token)
        _scan_status["last_result"] = result
    except Exception as e:
        _scan_status["last_error"] = str(e)
    finally:
        _scan_status["running"] = False

def _scan_github_sync(token: str):
    """Synchronous scan — runs in a separate thread via asyncio.to_thread."""
    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger("monitor")

    from github import Github
    g = Github(token)

    try:
        org = g.get_organization("tokamak-network")
    except Exception as e:
        raise RuntimeError("Failed to access org: {}".format(e))

    activities = []
    external_users = {}
    repos_scanned = 0

    def _add_activity(login, activity_type, repo_name, url, activity_date, details=""):
        if login in TEAM_MEMBERS:
            return
        dt_str = activity_date.isoformat() if activity_date else None
        activities.append((login, activity_type, repo_name, url, dt_str, details))
        if dt_str and (login not in external_users or dt_str > external_users[login]):
            external_users[login] = dt_str

    for repo in org.get_repos(sort="updated")[:30]:
        repos_scanned += 1
        repo_full = repo.full_name
        try:
            for sg in repo.get_stargazers_with_dates()[:50]:
                _add_activity(sg.user.login, "star", repo_full, "https://github.com/" + repo_full, sg.starred_at, "Starred " + repo_full)
        except Exception:
            pass
        try:
            for fork in repo.get_forks()[:20]:
                _add_activity(fork.owner.login, "fork", repo_full, fork.html_url, fork.created_at, "Forked " + repo_full)
        except Exception:
            pass
        try:
            for pr in repo.get_pulls(state="all", sort="updated", direction="desc")[:20]:
                _add_activity(pr.user.login, "pr", repo_full, pr.html_url, pr.created_at, pr.title)
        except Exception:
            pass
        try:
            for issue in repo.get_issues(state="all", sort="updated", direction="desc")[:20]:
                if not issue.pull_request:
                    _add_activity(issue.user.login, "issue", repo_full, issue.html_url, issue.created_at, issue.title)
        except Exception:
            pass

    # Save to DB synchronously
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")

    logger.info("Total activities collected: %d", len(activities))
    saved_acts = 0
    for login, atype, repo_name, url, dt_str, details in activities:
        try:
            conn.execute("""
                INSERT OR IGNORE INTO monitor_activities
                (github_username, activity_type, repo_name, activity_url, activity_date, details)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (login, atype, repo_name, url, dt_str, details))
            saved_acts += 1
        except Exception as e:
            logger.error("Failed to save activity: %s %s %s - %s", login, atype, url, e)
    conn.commit()
    logger.info("Activities saved: %d", saved_acts)

    # Analyze profiles synchronously
    analyzed = 0
    for username in list(external_users.keys())[:50]:
        try:
            user_obj = g.get_user(username)
            repos = list(user_obj.get_repos(sort="updated")[:10])
            langs = {}
            for r in repos:
                if r.language:
                    langs[r.language] = langs.get(r.language, 0) + 1

            interest = min(10, 3 + len([r for r in repos if r.language in {"Solidity", "TypeScript", "Rust"}]))
            tech_skill = min(10, 2 + user_obj.public_repos // 5 + user_obj.followers // 10)
            activity = min(10, 3 + len(repos) // 2)
            eco_rel = 5 if any(l in langs for l in ["Solidity", "Rust", "TypeScript"]) else 3

            scores = {
                "tokamak_interest": interest,
                "technical_skill": tech_skill,
                "activity_level": activity,
                "ecosystem_relevance": eco_rel,
            }

            last_active = external_users.get(username, datetime.utcnow().isoformat())
            recent_repos = [{"name": r.name, "language": r.language, "stars": r.stargazers_count} for r in repos]

            conn.execute("""
                INSERT INTO monitor_candidates (github_username, profile_url, bio, public_repos, followers, languages, contributions, scores, last_scanned)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(github_username) DO UPDATE SET
                    bio=excluded.bio, public_repos=excluded.public_repos, followers=excluded.followers,
                    languages=excluded.languages, scores=excluded.scores, last_scanned=excluded.last_scanned
            """, (
                username, f"https://github.com/{username}", user_obj.bio or "",
                user_obj.public_repos, user_obj.followers,
                json.dumps(langs), json.dumps(recent_repos),
                json.dumps(scores), last_active
            ))
            analyzed += 1
        except Exception as e:
            logger.error("Failed to analyze %s: %s", username, e)
            continue

    conn.commit()
    conn.close()
    return {"repos_scanned": repos_scanned, "external_users_found": len(external_users), "profiles_analyzed": analyzed}


@app.get("/api/monitor/candidates")
async def list_monitor_candidates(activity_within: str = ""):
    """List monitor candidates. activity_within: 1w, 1m, 3m to filter by last_scanned."""
    db = await get_db()
    query = "SELECT * FROM monitor_candidates"
    params = []  # type: list
    if activity_within in ("1w", "1m", "3m"):
        days_map = {"1w": 7, "1m": 30, "3m": 90}
        query += " WHERE last_scanned >= datetime('now', ?)"
        params.append("-{} days".format(days_map[activity_within]))
    query += " ORDER BY last_scanned DESC"
    rows = await db.execute(query, params)
    candidates = []
    for r in await rows.fetchall():
        c = dict(r)
        for f in ["languages", "contributions", "scores"]:
            if c.get(f):
                c[f] = json.loads(c[f])

        # Fetch recent activities
        act_rows = await db.execute(
            "SELECT activity_type, repo_name, activity_url, activity_date, details FROM monitor_activities WHERE github_username = ? ORDER BY activity_date DESC LIMIT 5",
            (c["github_username"],)
        )
        c["recent_activities"] = [dict(a) for a in await act_rows.fetchall()]

        # Activity type summary
        summary_rows = await db.execute(
            "SELECT activity_type, COUNT(*) as cnt FROM monitor_activities WHERE github_username = ? GROUP BY activity_type",
            (c["github_username"],)
        )
        c["activity_types"] = {row["activity_type"]: row["cnt"] for row in await summary_rows.fetchall()}

        candidates.append(c)
    await db.close()
    return candidates


@app.get("/api/monitor/candidates/{github_username}/activities")
async def get_monitor_activities(github_username: str):
    """Get all activities for a monitor candidate."""
    db = await get_db()
    rows = await db.execute(
        "SELECT activity_type, repo_name, activity_url, activity_date, details FROM monitor_activities WHERE github_username = ? ORDER BY activity_date DESC",
        (github_username,)
    )
    activities = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return activities


@app.get("/api/monitor/candidates/{github_username}")
async def get_monitor_candidate(github_username: str):
    db = await get_db()
    row = await db.execute("SELECT * FROM monitor_candidates WHERE github_username = ?", (github_username,))
    candidate = await row.fetchone()
    await db.close()
    if not candidate:
        raise HTTPException(404, "Candidate not found")
    c = dict(candidate)
    for f in ["languages", "contributions", "scores"]:
        if c.get(f):
            c[f] = json.loads(c[f])
    return c


# ── Team Profile Endpoints ────────────────────────────────────────────────


@app.post("/api/team/profile-scan")
async def team_profile_scan():
    """Trigger a full org scan and profile generation."""
    db = await get_db()
    try:
        result = await scan_org_profiles(db)
        if "error" in result:
            raise HTTPException(400, result["error"])
        return result
    finally:
        await db.close()


@app.get("/api/team/profiles")
async def list_team_profiles():
    """List all team member profiles with expertise."""
    db = await get_db()
    rows = await db.execute("SELECT * FROM team_profiles WHERE is_active = 1 ORDER BY review_count DESC, github_username")
    profiles = []
    for r in await rows.fetchall():
        p = dict(r)
        for f in ["expertise_areas", "top_repos", "languages"]:
            if p.get(f):
                p[f] = json.loads(p[f])
        profiles.append(p)
    await db.close()
    return profiles


@app.get("/api/team/profiles/{github_username}")
async def get_team_profile(github_username: str):
    """Get individual team member profile."""
    db = await get_db()
    row = await db.execute("SELECT * FROM team_profiles WHERE github_username = ?", (github_username,))
    profile = await row.fetchone()
    await db.close()
    if not profile:
        raise HTTPException(404, "Profile not found")
    p = dict(profile)
    for f in ["expertise_areas", "top_repos", "languages"]:
        if p.get(f):
            p[f] = json.loads(p[f])
    return p


@app.delete("/api/team/profiles/{github_username}")
async def delete_team_profile(github_username: str):
    """Delete a team member profile."""
    db = await get_db()
    row = await db.execute("SELECT id FROM team_profiles WHERE github_username = ?", (github_username,))
    if not await row.fetchone():
        await db.close()
        raise HTTPException(404, "Profile not found")
    await db.execute("DELETE FROM team_profiles WHERE github_username = ?", (github_username,))
    await db.commit()
    await db.close()
    return {"deleted": github_username}


class AddTeamMemberRequest(BaseModel):
    github_username: str


@app.post("/api/team/profiles")
async def add_team_profile(req: AddTeamMemberRequest):
    """Add a team member by GitHub username. Fetches profile from GitHub API and scans their activity."""
    import os
    from github import Github, GithubException
    from collections import Counter, defaultdict
    from datetime import datetime, timedelta
    from team_profiler import _repo_to_domains, _langs_to_expertise

    token = os.getenv("GITHUB_TOKEN", "")
    if not token:
        raise HTTPException(400, "GITHUB_TOKEN not configured")

    g = Github(token, per_page=100)
    username = req.github_username.strip().lstrip("@")

    # Fetch user info
    try:
        user = g.get_user(username)
    except GithubException:
        raise HTTPException(404, f"GitHub user '{username}' not found")

    # Check if already exists
    db = await get_db()
    row = await db.execute("SELECT id FROM team_profiles WHERE github_username = ?", (username,))
    if await row.fetchone():
        await db.close()
        raise HTTPException(409, f"'{username}' already exists in team profiles")

    # Scan tokamak-network org repos for this user's activity
    try:
        org = g.get_organization("tokamak-network")
        all_repos = list(org.get_repos(sort="updated", type="all"))[:50]
    except Exception as e:
        await db.close()
        raise HTTPException(500, f"Failed to access org repos: {e}")

    six_months_ago = datetime.utcnow() - timedelta(days=180)
    commits_per_repo = Counter()
    languages = Counter()
    domains = Counter()
    repos_detail = []

    for repo in all_repos:
        try:
            commits = list(repo.get_commits(author=username, since=six_months_ago))
            if not commits:
                continue
            count = len(commits)
            commits_per_repo[repo.name] = count

            repo_domains = _repo_to_domains(repo.name, getattr(repo, "topics", []) or [], repo.description)
            for d in repo_domains:
                domains[d] += count

            try:
                repo_langs = repo.get_languages()
                for lang, bytes_count in repo_langs.items():
                    languages[lang] += bytes_count
            except Exception:
                pass

            repos_detail.append({"name": repo.name, "commits": count, "language": max(repo.get_languages() or {"Unknown": 0}, key=lambda k: repo.get_languages().get(k, 0), default="Unknown")})
        except Exception:
            continue

    # Build expertise
    lang_expertise = _langs_to_expertise(dict(languages))
    domain_total = sum(domains.values()) or 1
    expertise = {**lang_expertise}
    for d, c in domains.items():
        expertise[d] = max(expertise.get(d, 0), min(c / domain_total, 1.0))

    review_count = sum(commits_per_repo.values())
    top_repos = sorted(repos_detail, key=lambda r: -r["commits"])[:5]
    top_languages = dict(Counter(languages).most_common(10))
    now = datetime.utcnow().isoformat()

    await db.execute("""
        INSERT INTO team_profiles (github_username, display_name, avatar_url, expertise_areas, top_repos, languages, review_count, last_active, last_profiled, is_active)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
    """, (
        username,
        user.name or username,
        user.avatar_url or "",
        json.dumps(expertise),
        json.dumps(top_repos),
        json.dumps(top_languages),
        review_count,
        now, now,
    ))
    await db.commit()
    await db.close()

    return {"added": username, "display_name": user.name or username, "repos_scanned": len(all_repos), "commits_found": review_count}


# ── LinkedIn Sourcing Endpoints ──────────────────────────────────────────────


class LinkedInSearchRequest(BaseModel):
    keywords: str = ""
    queries: list = []


@app.post("/api/linkedin/search")
async def linkedin_search(data: LinkedInSearchRequest):
    """Trigger candidate sourcing. Tries web search first, falls back to GitHub API."""
    result = await search_linkedin_candidates(
        keywords=data.keywords or None,
        queries=data.queries or None,
    )
    # If web search found nothing, fall back to GitHub API search
    if result.get("total_found", 0) == 0:
        result = await search_github_developers(
            keywords=data.keywords or None,
            max_per_query=15,
        )
    # Always attach saved_ids from DB if candidates were saved
    if result.get("total_saved", 0) > 0:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        rows = conn.execute(
            "SELECT id FROM linkedin_candidates WHERE created_at >= datetime('now', '-2 minutes') ORDER BY id DESC"
        ).fetchall()
        conn.close()
        result["saved_ids"] = [r[0] for r in rows]
    return result


@app.post("/api/github/search")
async def github_search(data: LinkedInSearchRequest):
    """Search GitHub directly for developer candidates."""
    result = await search_github_developers(
        keywords=data.keywords or None,
        queries=data.queries if data.queries else None,
        max_per_query=20,
    )
    return result


@app.get("/api/linkedin/candidates")
async def linkedin_candidates(status: str = "", limit: int = 100, offset: int = 0):
    """List LinkedIn candidates with scores and total count."""
    candidates = get_linkedin_candidates(
        status=status or None,
        limit=limit,
        offset=offset,
    )
    # Get total count for pagination
    db = await get_db()
    if status:
        row = await db.execute("SELECT COUNT(*) as cnt FROM linkedin_candidates WHERE status=?", (status,))
    else:
        row = await db.execute("SELECT COUNT(*) as cnt FROM linkedin_candidates")
    total = (await row.fetchone())["cnt"]
    await db.close()
    return {"candidates": candidates, "total": total}


class OutreachRequest(BaseModel):
    status: str = "outreach"
    notes: str = ""
    template_id: str = ""
    message_sent: str = ""
    channel: str = "linkedin_dm"


@app.get("/api/templates/outreach")
async def get_outreach_templates():
    """Parse outreach_templates.md and return structured JSON."""
    import re
    templates_path = os.path.join(os.path.dirname(__file__), "..", "templates", "outreach_templates.md")
    if not os.path.exists(templates_path):
        raise HTTPException(404, "Templates file not found")
    with open(templates_path, "r", encoding="utf-8") as f:
        content = f.read()

    templates = []
    # Split by ## N. Title — produces ['preamble', '1', 'Title1', 'content1', '2', 'Title2', 'content2', ...]
    sections = re.split(r'^## (\d+)\.\s+(.+)$', content, flags=re.MULTILINE)
    i = 1
    while i + 2 < len(sections):
        tid = sections[i].strip()
        name = sections[i + 1].strip()
        rest = sections[i + 2]

        # Parse English and Korean blocks
        langs = {}
        subjects = {}
        for lang_label, lang_key in [("### English", "en"), ("### 한국어", "kr")]:
            if lang_label in rest:
                after = rest.split(lang_label, 1)[1]
                # Extract subject line if present
                subj_match = re.search(r'\*\*Subject:\*\*\s*(.+)', after)
                if subj_match:
                    subjects[lang_key] = subj_match.group(1).strip()
                code_match = re.search(r'```\n?(.*?)```', after, re.DOTALL)
                if code_match:
                    langs[lang_key] = code_match.group(1).strip()

        # If no language subsections, grab first code block as English
        if not langs:
            code_match = re.search(r'```\n?(.*?)```', rest, re.DOTALL)
            if code_match:
                langs["en"] = code_match.group(1).strip()

        # Extract variables from template text
        all_text = " ".join(langs.values())
        variables = sorted(set(re.findall(r'\{([^}]+)\}', all_text)))

        for lang_key, body in langs.items():
            templates.append({
                "id": "{}_{}".format(tid, lang_key),
                "name": name,
                "language": lang_key,
                "subject": subjects.get(lang_key, ""),
                "body": body,
                "variables": variables,
            })

        i += 3

    return templates


@app.post("/api/linkedin/candidates/{candidate_id}/outreach")
async def linkedin_outreach(candidate_id: int, data: OutreachRequest):
    """Mark candidate for outreach (status update) and optionally save outreach history."""
    success = update_linkedin_status(candidate_id, data.status, data.notes)
    if not success:
        raise HTTPException(404, "Candidate not found")

    # Save to outreach_history if message was sent
    if data.message_sent:
        db = await get_db()
        await db.execute(
            "INSERT INTO outreach_history (candidate_id, candidate_type, template_used, message_sent, channel, status, sent_by) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (candidate_id, "linkedin", data.template_id, data.message_sent, data.channel, "sent", data.notes or "")
        )
        await db.commit()
        await db.close()

    return {"id": candidate_id, "status": data.status}


@app.get("/api/linkedin/candidates/{candidate_id}/outreach-history")
async def get_outreach_history(candidate_id: int):
    """Returns outreach history for a candidate."""
    db = await get_db()
    rows = await db.execute(
        "SELECT * FROM outreach_history WHERE candidate_id = ? ORDER BY sent_at DESC",
        (candidate_id,)
    )
    history = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return history


# ── Monitor LinkedIn Lookup ───────────────────────────────────────────────


async def _find_linkedin_for_monitor_candidate(username: str, db) -> Optional[str]:
    """Find LinkedIn URL for a single monitor candidate. Returns URL or None."""
    import httpx
    import re

    token = os.getenv("GITHUB_TOKEN", "")
    headers = {}
    if token:
        headers["Authorization"] = "token {}".format(token)

    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                "https://api.github.com/users/{}".format(username),
                headers=headers,
            )
            if resp.status_code != 200:
                return None
            user_data = resp.json()
    except Exception:
        return None

    real_name = user_data.get("name", "") or ""
    location = user_data.get("location", "") or ""
    bio = user_data.get("bio", "") or ""
    blog = user_data.get("blog", "") or ""

    # 1. Check profile fields for LinkedIn URL
    for field in [blog, bio]:
        match = re.search(r'linkedin\.com/in/([a-zA-Z0-9_-]+)', field)
        if match:
            url = "https://www.linkedin.com/in/{}".format(match.group(1))
            await db.execute("UPDATE monitor_candidates SET linkedin_url = ? WHERE github_username = ?", (url, username))
            await db.commit()
            return url

    # 2. Check GitHub social accounts API
    try:
        async with httpx.AsyncClient(timeout=15) as client:
            resp = await client.get(
                "https://api.github.com/users/{}/social_accounts".format(username),
                headers=headers,
            )
            if resp.status_code == 200:
                for account in resp.json():
                    if "linkedin" in account.get("url", ""):
                        url = account["url"]
                        if not url.startswith("https://"):
                            url = "https://" + url.lstrip("/")
                        await db.execute("UPDATE monitor_candidates SET linkedin_url = ? WHERE github_username = ?", (url, username))
                        await db.commit()
                        return url
    except Exception:
        pass

    if not real_name:
        return None

    # Use existing search function
    from github_linkedin import find_linkedin_for_github_user
    result = await find_linkedin_for_github_user(real_name, location, bio)
    if result:
        url = result["profile_url"]
        await db.execute("UPDATE monitor_candidates SET linkedin_url = ? WHERE github_username = ?", (url, username))
        await db.commit()
        return url

    return None


@app.post("/api/monitor/find-linkedin")
async def monitor_find_linkedin_all():
    """Find LinkedIn profiles for all monitor candidates missing linkedin_url."""
    db = await get_db()
    rows = await db.execute(
        "SELECT github_username FROM monitor_candidates WHERE linkedin_url IS NULL OR linkedin_url = ''"
    )
    candidates = [dict(r) for r in await rows.fetchall()]

    found_list = []
    for c in candidates:
        username = c["github_username"]
        url = await _find_linkedin_for_monitor_candidate(username, db)
        if url:
            found_list.append({"username": username, "linkedin_url": url})

    await db.close()
    return {"checked": len(candidates), "found": len(found_list), "candidates": found_list}


@app.post("/api/monitor/find-linkedin/{username}")
async def monitor_find_linkedin_single(username: str):
    """Find LinkedIn profile for a single monitor candidate."""
    db = await get_db()
    row = await db.execute("SELECT id FROM monitor_candidates WHERE github_username = ?", (username,))
    if not await row.fetchone():
        await db.close()
        raise HTTPException(404, "Candidate not found")

    url = await _find_linkedin_for_monitor_candidate(username, db)
    await db.close()
    if url:
        return {"username": username, "linkedin_url": url}
    return {"username": username, "linkedin_url": None, "message": "LinkedIn profile not found"}


@app.post("/api/linkedin/bridge")
async def linkedin_bridge():
    """Find LinkedIn profiles for GitHub monitor candidates."""
    result = await bridge_github_candidates()
    return result


@app.get("/api/candidates/{candidate_id}/match")
async def get_candidate_match(candidate_id: int):
    """Get team matching scores for a candidate."""
    result = match_candidate_to_team(candidate_id)
    if "error" in result:
        raise HTTPException(status_code=404, detail=result["error"])
    return result


# ══════════════════════════════════════════════════════════════════════════════
# HR / Payroll Endpoints
# ══════════════════════════════════════════════════════════════════════════════

from tax_calculator import simulate_annual_tax, monthly_tax_burden, calculate_tax
import calendar as _calendar
from pydantic import BaseModel as _BM


class HRMemberCreate(BaseModel):
    name: str
    github: str = ""
    role: str
    monthly_usdt: float
    wallet_address: str = ""
    contract_start: str = ""

class HRMemberUpdate(BaseModel):
    name: Optional[str] = None
    github: Optional[str] = None
    role: Optional[str] = None
    monthly_usdt: Optional[float] = None
    wallet_address: Optional[str] = None
    contract_start: Optional[str] = None
    contract_end: Optional[str] = None
    is_active: Optional[int] = None
    name_kr: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    personal_email: Optional[str] = None
    birthday: Optional[str] = None
    education: Optional[str] = None
    nationality: Optional[str] = None
    is_rnd: Optional[str] = None
    address: Optional[str] = None
    company: Optional[str] = None

class PayrollConfirm(BaseModel):
    year: int
    month: int

class PayrollCreate(BaseModel):
    member_id: int
    year: int
    month: int
    usdt_amount: float
    krw_rate: float
    krw_amount: float
    tax_simulated: float
    net_pay_krw: float
    status: str = "paid"

class PayrollUpdate(BaseModel):
    usdt_amount: Optional[float] = None
    krw_rate: Optional[float] = None
    krw_amount: Optional[float] = None
    tax_simulated: Optional[float] = None
    net_pay_krw: Optional[float] = None
    status: Optional[str] = None


# ── HR Members ──

@app.get("/api/hr/members/download")
async def download_members(active: Optional[int] = None):
    """팀원 목록 엑셀 다운로드"""
    import openpyxl, io
    from fastapi.responses import StreamingResponse

    db = await get_db()
    if active is not None:
        rows = await db.execute("SELECT * FROM hr_members WHERE is_active=? ORDER BY id", (active,))
    else:
        rows = await db.execute("SELECT * FROM hr_members ORDER BY id")
    members = [dict(r) for r in await rows.fetchall()]
    await db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "팀원"
    ws.append(["이름", "GitHub", "직책", "월급(USDT)", "지갑주소", "계약시작일", "퇴직일", "상태"])
    for m in members:
        ws.append([m["name"], m["github"], m["role"], m["monthly_usdt"], m["wallet_address"], m["contract_start"], m.get("contract_end", ""), "재직" if m["is_active"] else "퇴직"])
    for col in ["A","B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    label = "active" if active == 1 else "retired" if active == 0 else "all"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=members_{label}.xlsx"})


@app.post("/api/hr/members/upload")
async def upload_members(file: UploadFile = File(...)):
    """팀원 엑셀 업로드 (신규 추가 또는 기존 업데이트)"""
    import openpyxl, io

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    db = await get_db()
    added, updated = 0, 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        github = str(row[1] or "").strip()
        role = str(row[2] or "").strip()
        monthly_usdt = float(row[3] or 0)
        wallet = str(row[4] or "").strip()
        contract_start = str(row[5] or "").strip()
        contract_end = str(row[6] or "").strip() if len(row) > 6 else ""
        status = str(row[7] or "재직").strip() if len(row) > 7 else "재직"
        is_active = 0 if status == "퇴직" else 1

        existing = await db.execute("SELECT id FROM hr_members WHERE name=?", (name,))
        ex = await existing.fetchone()
        if ex:
            await db.execute(
                "UPDATE hr_members SET github=?, role=?, monthly_usdt=?, wallet_address=?, contract_start=?, contract_end=?, is_active=? WHERE id=?",
                (github, role, monthly_usdt, wallet, contract_start, contract_end or None, is_active, ex["id"]))
            updated += 1
        else:
            await db.execute(
                "INSERT INTO hr_members (name, github, role, monthly_usdt, wallet_address, contract_start, contract_end, is_active) VALUES (?,?,?,?,?,?,?,?)",
                (name, github, role, monthly_usdt, wallet, contract_start, contract_end or None, is_active))
            added += 1

    await db.commit()
    await db.close()
    return {"added": added, "updated": updated, "message": f"{added}명 추가, {updated}명 업데이트"}


@app.get("/api/hr/members")
async def list_hr_members(active: Optional[int] = None):
    db = await get_db()
    if active is not None:
        rows = await db.execute("SELECT * FROM hr_members WHERE is_active=? ORDER BY id", (active,))
    else:
        rows = await db.execute("SELECT * FROM hr_members WHERE is_active=1 ORDER BY id")
    result = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return result

@app.get("/api/hr/members/{member_id}")
async def get_hr_member(member_id: int):
    db = await get_db()
    row = await db.execute("SELECT * FROM hr_members WHERE id=?", (member_id,))
    member = await row.fetchone()
    if not member:
        await db.close()
        raise HTTPException(404, "Member not found")
    m = dict(member)
    payrolls = await db.execute("SELECT * FROM payrolls WHERE member_id=? ORDER BY year DESC, month DESC", (member_id,))
    m["payrolls"] = [dict(p) for p in await payrolls.fetchall()]
    incentives = await db.execute("SELECT * FROM incentives WHERE member_id=? ORDER BY year DESC, quarter DESC", (member_id,))
    m["incentives"] = [dict(i) for i in await incentives.fetchall()]
    await db.close()
    return m

@app.post("/api/hr/members")
async def create_hr_member(data: HRMemberCreate):
    db = await get_db()
    cursor = await db.execute(
        "INSERT INTO hr_members (name, github, role, monthly_usdt, wallet_address, contract_start) VALUES (?,?,?,?,?,?)",
        (data.name, data.github, data.role, data.monthly_usdt, data.wallet_address, data.contract_start))
    await db.commit()
    mid = cursor.lastrowid
    await db.close()
    return {"id": mid, "message": "Member created"}

@app.put("/api/hr/members/{member_id}")
async def update_hr_member(member_id: int, data: HRMemberUpdate):
    db = await get_db()
    row = await db.execute("SELECT * FROM hr_members WHERE id=?", (member_id,))
    if not await row.fetchone():
        await db.close()
        raise HTTPException(404, "Member not found")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if updates:
        set_clause = ", ".join(f"{k}=?" for k in updates)
        await db.execute(f"UPDATE hr_members SET {set_clause} WHERE id=?", list(updates.values()) + [member_id])
        await db.commit()
    await db.close()
    return {"message": "Updated"}

@app.delete("/api/hr/members/{member_id}")
async def delete_hr_member(member_id: int, permanent: bool = False):
    db = await get_db()
    if permanent:
        await db.execute("DELETE FROM payrolls WHERE member_id=?", (member_id,))
        await db.execute("DELETE FROM incentives WHERE member_id=?", (member_id,))
        await db.execute("DELETE FROM reserves WHERE member_id=?", (member_id,))
        await db.execute("DELETE FROM hr_members WHERE id=?", (member_id,))
    else:
        await db.execute("UPDATE hr_members SET is_active=0 WHERE id=?", (member_id,))
    await db.commit()
    await db.close()
    return {"message": "Deleted" if permanent else "Deactivated"}

@app.post("/api/hr/members/{member_id}/retire")
async def retire_hr_member(member_id: int, data: dict):
    """퇴직 처리 (is_active=0, contract_end 기록)"""
    db = await get_db()
    contract_end = data.get("contract_end", datetime.now().strftime("%Y-%m-%d"))
    await db.execute("UPDATE hr_members SET is_active=0, contract_end=? WHERE id=?", (contract_end, member_id))
    await db.commit()
    await db.close()
    return {"message": "퇴직 처리 완료"}

@app.post("/api/hr/members/{member_id}/reinstate")
async def reinstate_hr_member(member_id: int):
    """복직 처리 (is_active=1, contract_end 제거)"""
    db = await get_db()
    await db.execute("UPDATE hr_members SET is_active=1, contract_end=NULL WHERE id=?", (member_id,))
    await db.commit()
    await db.close()
    return {"message": "복직 처리 완료"}


# ── Payroll ──

@app.get("/api/hr/payroll/download")
async def download_payroll(year: int = 2026, month: Optional[int] = None):
    """급여 월별 엑셀 다운로드"""
    import openpyxl, io
    from fastapi.responses import StreamingResponse

    db = await get_db()
    if month:
        rows = await db.execute("""
            SELECT p.*, m.name, m.role FROM payrolls p
            JOIN hr_members m ON p.member_id = m.id
            WHERE p.year=? AND p.month=? ORDER BY m.name
        """, (year, month))
    else:
        rows = await db.execute("""
            SELECT p.*, m.name, m.role FROM payrolls p
            JOIN hr_members m ON p.member_id = m.id
            WHERE p.year=? ORDER BY p.month, m.name
        """, (year,))
    data = [dict(r) for r in await rows.fetchall()]
    await db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년 급여"
    ws.append(["이름", "직책", "연도", "월", "USDT", "환율", "KRW", "세금(KRW)", "실지급(KRW)", "상태"])
    for p in data:
        ws.append([p["name"], p["role"], p["year"], p["month"], p["usdt_amount"], p["krw_rate"], p["krw_amount"], p["tax_simulated"], p["net_pay_krw"], p["status"]])
    for col in ["A","B","C","D","E","F","G","H","I","J"]:
        ws.column_dimensions[col].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"payroll_{year}_{month}m.xlsx" if month else f"payroll_{year}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})

@app.get("/api/hr/payroll")
async def list_payroll(year: int = 2026, month: Optional[int] = None):
    db = await get_db()
    if month:
        rows = await db.execute("""
            SELECT p.*, m.name, m.role, m.wallet_address FROM payrolls p
            JOIN hr_members m ON p.member_id = m.id
            WHERE p.year=? AND p.month=? ORDER BY m.name
        """, (year, month))
    else:
        rows = await db.execute("""
            SELECT p.*, m.name, m.role, m.wallet_address FROM payrolls p
            JOIN hr_members m ON p.member_id = m.id
            WHERE p.year=? ORDER BY p.month DESC, m.name
        """, (year,))
    result = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return result

@app.post("/api/hr/payroll/add")
async def add_payroll(data: PayrollCreate):
    db = await get_db()
    existing = await db.execute(
        "SELECT id FROM payrolls WHERE member_id=? AND year=? AND month=?",
        (data.member_id, data.year, data.month))
    if await existing.fetchone():
        await db.close()
        raise HTTPException(400, "해당 월의 급여 이력이 이미 존재합니다")
    cursor = await db.execute(
        "INSERT INTO payrolls (member_id, year, month, usdt_amount, krw_rate, krw_amount, tax_simulated, reserve_tokamak, net_pay_krw, status) VALUES (?,?,?,?,?,?,?,0,?,?)",
        (data.member_id, data.year, data.month, data.usdt_amount, data.krw_rate, data.krw_amount, data.tax_simulated, data.net_pay_krw, data.status))
    await db.commit()
    pid = cursor.lastrowid
    await db.close()
    return {"id": pid, "message": "Created"}

@app.put("/api/hr/payroll/{payroll_id}")
async def update_payroll(payroll_id: int, data: PayrollUpdate):
    db = await get_db()
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if updates:
        set_clause = ", ".join(f"{k}=?" for k in updates)
        await db.execute(f"UPDATE payrolls SET {set_clause} WHERE id=?", list(updates.values()) + [payroll_id])
        await db.commit()
    await db.close()
    return {"message": "Updated"}

@app.delete("/api/hr/payroll/{payroll_id}")
async def delete_payroll(payroll_id: int):
    db = await get_db()
    await db.execute("DELETE FROM payrolls WHERE id=?", (payroll_id,))
    await db.commit()
    await db.close()
    return {"message": "Deleted"}

@app.post("/api/hr/payroll/confirm")
async def confirm_payroll(data: PayrollConfirm):
    db = await get_db()
    await db.execute("UPDATE payrolls SET status='confirmed', confirmed_at=datetime('now') WHERE year=? AND month=? AND status='estimated'", (data.year, data.month))
    await db.commit()
    await db.close()
    return {"message": "Confirmed"}

@app.post("/api/hr/payroll/pay")
async def pay_payroll(data: PayrollConfirm):
    db = await get_db()
    await db.execute("UPDATE payrolls SET status='paid', paid_at=datetime('now') WHERE year=? AND month=? AND status='confirmed'", (data.year, data.month))
    await db.commit()
    await db.close()
    return {"message": "Marked as paid"}


# ── Dashboard ──

@app.get("/api/hr/dashboard")
async def hr_dashboard():
    import calendar as _cal
    import math
    from datetime import date

    db = await get_db()
    today = date.today()
    year, month = today.year, today.month

    # 이번 달 급여 데이터
    rows = await db.execute("SELECT * FROM payrolls WHERE year=? AND month=?", (year, month))
    payrolls = [dict(r) for r in await rows.fetchall()]
    total_usdt = sum(p["usdt_amount"] for p in payrolls)
    total_krw = sum(p["krw_amount"] for p in payrolls)
    total_tax = sum(p["tax_simulated"] for p in payrolls)

    # 데이터 없으면 재직 팀원 기준 예상치
    if not payrolls:
        m_rows = await db.execute("SELECT monthly_usdt FROM hr_members WHERE is_active=1")
        members = [dict(r) for r in await m_rows.fetchall()]
        total_usdt = sum(m["monthly_usdt"] for m in members)
        member_count = len(members)
    else:
        member_count = len(payrolls)

    # 최근 트랜잭션
    tx_rows = await db.execute("SELECT * FROM hr_transactions ORDER BY timestamp DESC LIMIT 5")
    txs = [dict(r) for r in await tx_rows.fetchall()]

    # 급여일 (이번 달 마지막 영업일)
    last_day = date(today.year, today.month, _cal.monthrange(today.year, today.month)[1])
    while last_day.weekday() >= 5:
        last_day = last_day.replace(day=last_day.day - 1)
    d_day = max((last_day - today).days, 0)

    # 연간 세금 누적 적립금
    tax_acc_row = await db.execute("SELECT SUM(tax_simulated) as total_tax_yr FROM payrolls WHERE year=? AND tax_simulated > 0", (year,))
    tax_acc = await tax_acc_row.fetchone()
    total_tax_year_krw = tax_acc["total_tax_yr"] or 0

    rate_row = await db.execute("SELECT AVG(krw_rate) as avg_rate FROM payrolls WHERE year=? AND krw_rate > 0", (year,))
    rate = await rate_row.fetchone()
    avg_rate = rate["avg_rate"] or 1
    total_tax_year_usdt = math.ceil(total_tax_year_krw / avg_rate / 10) * 10 if total_tax_year_krw > 0 else 0

    # 연간 총 지급 USDT
    annual_row = await db.execute("SELECT SUM(usdt_amount) as total FROM payrolls WHERE year=?", (year,))
    annual = await annual_row.fetchone()
    annual_usdt = annual["total"] or 0

    await db.close()
    return {
        "current_month": {"year": year, "month": month, "total_usdt": round(total_usdt, 2), "total_krw": round(total_krw), "total_tax": round(total_tax), "member_count": member_count},
        "recent_transactions": txs,
        "d_day": d_day,
        "payday": last_day.isoformat(),
        "annual_usdt": round(annual_usdt, 2),
        "reserves": {"total_tax_usdt": total_tax_year_usdt, "total_tax_krw": round(total_tax_year_krw)},
    }


# ── Settings & Wallets ──

@app.get("/api/hr/settings")
async def get_settings():
    db = await get_db()
    rows = await db.execute("SELECT key, value FROM hr_settings")
    settings = {r["key"]: r["value"] for r in await rows.fetchall()}
    await db.close()
    return settings

@app.put("/api/hr/settings")
async def update_settings(data: dict):
    db = await get_db()
    for k, v in data.items():
        await db.execute("INSERT OR REPLACE INTO hr_settings (key, value) VALUES (?,?)", (k, v))
    await db.commit()
    await db.close()
    return {"message": "Updated"}

@app.get("/api/hr/wallets")
async def list_wallets():
    db = await get_db()
    rows = await db.execute("SELECT * FROM hr_wallets WHERE is_active=1 ORDER BY id")
    result = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return result

class WalletCreate(BaseModel):
    label: str
    address: str
    chain: str = "ERC-20"

@app.post("/api/hr/wallets")
async def add_wallet(data: WalletCreate):
    db = await get_db()
    cursor = await db.execute(
        "INSERT INTO hr_wallets (label, address, chain, created_at) VALUES (?,?,?,datetime('now'))",
        (data.label, data.address, data.chain))
    await db.commit()
    wid = cursor.lastrowid
    await db.close()
    return {"id": wid, "message": "Added"}

@app.put("/api/hr/wallets/{wallet_id}")
async def update_wallet(wallet_id: int, data: WalletCreate):
    db = await get_db()
    await db.execute("UPDATE hr_wallets SET label=?, address=?, chain=? WHERE id=?",
                     (data.label, data.address, data.chain, wallet_id))
    await db.commit()
    await db.close()
    return {"message": "Updated"}

@app.delete("/api/hr/wallets/{wallet_id}")
async def delete_wallet(wallet_id: int):
    db = await get_db()
    await db.execute("DELETE FROM hr_wallets WHERE id=?", (wallet_id,))
    await db.commit()
    await db.close()
    return {"message": "Deleted"}


# ── Expenses (경비 정산) ──

class ExpenseCreate(BaseModel):
    member_id: int
    year: int
    month: int
    amount_usdt: float
    category: str
    description: str = ""
    tx_hash: str = ""
    status: str = "pending"
    expense_date: str = ""

class ExpenseUpdate(BaseModel):
    amount_usdt: Optional[float] = None
    category: Optional[str] = None
    description: Optional[str] = None
    tx_hash: Optional[str] = None
    status: Optional[str] = None
    expense_date: Optional[str] = None

@app.get("/api/hr/expenses")
async def list_expenses(year: int = 2026, month: Optional[int] = None):
    db = await get_db()
    if month:
        rows = await db.execute("""
            SELECT e.*, m.name, m.role FROM expenses e
            JOIN hr_members m ON e.member_id = m.id
            WHERE e.year=? AND e.month=? ORDER BY e.expense_date DESC, m.name
        """, (year, month))
    else:
        rows = await db.execute("""
            SELECT e.*, m.name, m.role FROM expenses e
            JOIN hr_members m ON e.member_id = m.id
            WHERE e.year=? ORDER BY e.month DESC, e.expense_date DESC, m.name
        """, (year,))
    result = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return result

@app.post("/api/hr/expenses")
async def create_expense(data: ExpenseCreate):
    db = await get_db()
    cursor = await db.execute(
        "INSERT INTO expenses (member_id, year, month, amount_usdt, category, description, tx_hash, status, expense_date, created_at) VALUES (?,?,?,?,?,?,?,?,?,datetime('now'))",
        (data.member_id, data.year, data.month, data.amount_usdt, data.category, data.description, data.tx_hash, data.status, data.expense_date))
    await db.commit()
    eid = cursor.lastrowid
    await db.close()
    return {"id": eid, "message": "Created"}

@app.put("/api/hr/expenses/{expense_id}")
async def update_expense(expense_id: int, data: ExpenseUpdate):
    db = await get_db()
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if updates:
        set_clause = ", ".join(f"{k}=?" for k in updates)
        await db.execute(f"UPDATE expenses SET {set_clause} WHERE id=?", list(updates.values()) + [expense_id])
        await db.commit()
    await db.close()
    return {"message": "Updated"}

@app.delete("/api/hr/expenses/{expense_id}")
async def delete_expense(expense_id: int):
    db = await get_db()
    await db.execute("DELETE FROM expenses WHERE id=?", (expense_id,))
    await db.commit()
    await db.close()
    return {"message": "Deleted"}

@app.get("/api/hr/expenses/download")
async def download_expenses(year: int = 2026, month: Optional[int] = None):
    import openpyxl, io
    from fastapi.responses import StreamingResponse

    db = await get_db()
    if month:
        rows = await db.execute("""
            SELECT e.*, m.name FROM expenses e JOIN hr_members m ON e.member_id = m.id
            WHERE e.year=? AND e.month=? ORDER BY e.expense_date, m.name
        """, (year, month))
    else:
        rows = await db.execute("""
            SELECT e.*, m.name FROM expenses e JOIN hr_members m ON e.member_id = m.id
            WHERE e.year=? ORDER BY e.month, e.expense_date, m.name
        """, (year,))
    data = [dict(r) for r in await rows.fetchall()]
    await db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년 경비"
    ws.append(["이름", "연도", "월", "금액(USDT)", "카테고리", "내용", "TX Hash", "상태", "발생일"])
    for e in data:
        ws.append([e["name"], e["year"], e["month"], e["amount_usdt"], e["category"], e["description"], e["tx_hash"], e["status"], e["expense_date"]])
    for col in ["A","B","C","D","E","F","G","H","I"]:
        ws.column_dimensions[col].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"expenses_{year}_{month}m.xlsx" if month else f"expenses_{year}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ── Fiat Transactions (법인 입출금) ──

@app.get("/api/hr/fiat")
async def list_fiat(currency: str = "", direction: str = "", source: str = "", year: Optional[int] = None, month: Optional[int] = None, limit: int = 100, offset: int = 0):
    db = await get_db()
    conditions, params = [], []
    if currency:
        conditions.append("currency=?")
        params.append(currency)
    if direction:
        conditions.append("direction=?")
        params.append(direction)
    if source:
        conditions.append("source=?")
        params.append(source)
    if year:
        conditions.append("CAST(substr(tx_date, 1, 4) AS INTEGER)=?")
        params.append(year)
    if month:
        conditions.append("CAST(substr(tx_date, 6, 2) AS INTEGER)=?")
        params.append(month)
    where = " WHERE " + " AND ".join(conditions) if conditions else ""
    rows = await db.execute(f"SELECT * FROM fiat_transactions{where} ORDER BY tx_date DESC LIMIT ? OFFSET ?", params + [limit, offset])
    data = [dict(r) for r in await rows.fetchall()]
    count_row = await db.execute(f"SELECT COUNT(*) as cnt FROM fiat_transactions{where}", params)
    total = (await count_row.fetchone())["cnt"]
    await db.close()
    return {"transactions": data, "total": total}

@app.get("/api/hr/fiat/summary")
async def fiat_summary(year: Optional[int] = None, month: Optional[int] = None, source: str = ""):
    db = await get_db()
    conditions = ["(status='COMPLETED' OR status='completed')"]
    params: list = []
    if year:
        conditions.append("CAST(substr(tx_date, 1, 4) AS INTEGER)=?")
        params.append(year)
    if month:
        conditions.append("CAST(substr(tx_date, 6, 2) AS INTEGER)=?")
        params.append(month)
    if source:
        conditions.append("source=?")
        params.append(source)
    where = " WHERE " + " AND ".join(conditions)
    rows = await db.execute(f"""
        SELECT currency, direction,
            COUNT(*) as count, SUM(amount) as total_amount
        FROM fiat_transactions{where}
        GROUP BY currency, direction ORDER BY currency, direction
    """, params)
    data = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return data

@app.post("/api/hr/fiat/upload-wise")
async def upload_wise_csv(file: UploadFile = File(...)):
    """WISE CSV 업로드"""
    import csv, io

    content = (await file.read()).decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))

    db = await get_db()
    existing = set()
    rows = await db.execute("SELECT tx_id FROM fiat_transactions WHERE source='WISE'")
    for r in await rows.fetchall():
        existing.add(r["tx_id"])

    added = 0
    for row in reader:
        tx_id = row.get("ID", "").strip()
        if not tx_id or tx_id in existing:
            continue
        status = row.get("Status", "")
        direction = row.get("Direction", "")
        amount = float(row.get("Source amount (after fees)") or 0)
        currency = row.get("Source currency", "") or row.get("Source fee currency", "")
        counterparty = row.get("Target name", "") or row.get("Source name", "")
        tx_date = row.get("Finished on", "") or row.get("Created on", "")

        await db.execute(
            "INSERT INTO fiat_transactions (tx_id, source, direction, status, amount, currency, counterparty, category, reference, note, exchange_rate, tx_date, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))",
            (tx_id, "WISE", direction, status, amount, currency, counterparty,
             row.get("Category", ""), row.get("Reference", ""), row.get("Note", ""),
             float(row.get("Exchange rate") or 0), tx_date))
        added += 1

    await db.commit()
    await db.close()
    return {"added": added, "message": f"WISE: {added}건 업로드 완료"}

@app.post("/api/hr/fiat/upload-aspire")
async def upload_aspire_excel(file: UploadFile = File(...)):
    """Aspire 엑셀 업로드 (USD/SGD/GBP)"""
    import openpyxl, io

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    # Detect currency from header (e.g., "Amount (USD)")
    headers = [str(c.value or "") for c in ws[1]]
    currency = "USD"
    for h in headers:
        if "Amount" in h and "(" in h:
            currency = h.split("(")[1].replace(")", "").strip()
            break

    db = await get_db()
    existing = set()
    rows = await db.execute("SELECT tx_id FROM fiat_transactions WHERE source='Aspire'")
    for r in await rows.fetchall():
        existing.add(r["tx_id"])

    added = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        tx_id = str(row[1]).strip()
        if tx_id in existing:
            continue
        tx_type = str(row[2] or "")
        amount = float(row[3] or 0)
        counterparty = str(row[4] or "")
        reference = str(row[8] or "")
        category = str(row[10] or "")
        note = str(row[11] or "")
        balance = float(row[12] or 0) if row[12] else 0
        tx_date = str(row[16] or "")
        direction = "IN" if tx_type == "credit" else "OUT"

        await db.execute(
            "INSERT INTO fiat_transactions (tx_id, source, direction, status, amount, currency, counterparty, category, reference, note, exchange_rate, balance, tx_date, created_at) VALUES (?,?,?,?,?,?,?,?,?,?,0,?,?,datetime('now'))",
            (tx_id, "Aspire", direction, "COMPLETED", abs(amount), currency, counterparty,
             category, reference, note, balance, tx_date))
        added += 1

    await db.commit()
    await db.close()
    return {"added": added, "message": f"Aspire ({currency}): {added}건 업로드 완료"}

@app.delete("/api/hr/fiat/{tx_id}")
async def delete_fiat(tx_id: int):
    db = await get_db()
    await db.execute("DELETE FROM fiat_transactions WHERE id=?", (tx_id,))
    await db.commit()
    await db.close()
    return {"message": "Deleted"}


# ── Member Wallets ──

@app.get("/api/hr/members/{member_id}/wallets")
async def list_member_wallets(member_id: int):
    db = await get_db()
    rows = await db.execute("SELECT * FROM member_wallets WHERE member_id=? ORDER BY id", (member_id,))
    result = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return result

@app.post("/api/hr/members/{member_id}/wallets")
async def add_member_wallet(member_id: int, data: WalletCreate):
    db = await get_db()
    cursor = await db.execute(
        "INSERT INTO member_wallets (member_id, label, address, chain, created_at) VALUES (?,?,?,?,datetime('now'))",
        (member_id, data.label, data.address, data.chain))
    await db.commit()
    wid = cursor.lastrowid
    await db.close()
    return {"id": wid, "message": "Added"}

@app.put("/api/hr/members/wallets/{wallet_id}")
async def update_member_wallet(wallet_id: int, data: WalletCreate):
    db = await get_db()
    await db.execute("UPDATE member_wallets SET label=?, address=?, chain=? WHERE id=?",
                     (data.label, data.address, data.chain, wallet_id))
    await db.commit()
    await db.close()
    return {"message": "Updated"}

@app.delete("/api/hr/members/wallets/{wallet_id}")
async def delete_member_wallet(wallet_id: int):
    db = await get_db()
    await db.execute("DELETE FROM member_wallets WHERE id=?", (wallet_id,))
    await db.commit()
    await db.close()
    return {"message": "Deleted"}

@app.get("/api/hr/address-map")
async def get_address_map():
    """모든 지갑 주소 → 이름/라벨 매핑 (트랜잭션 표시용)"""
    db = await get_db()
    addr_map = {}

    # 관리자 지갑 (설정)
    rows = await db.execute("SELECT label, address FROM hr_wallets WHERE is_active=1")
    for r in await rows.fetchall():
        addr_map[r["address"].lower()] = r["label"]

    # 팀원 지갑
    rows = await db.execute("""
        SELECT mw.label, mw.address, m.name FROM member_wallets mw
        JOIN hr_members m ON mw.member_id = m.id
    """)
    for r in await rows.fetchall():
        addr_map[r["address"].lower()] = f"{r['name']} ({r['label']})"

    await db.close()
    return addr_map


# ── Etherscan Transaction Sync ──

USDT_CONTRACT = "0xdAC17F958D2ee523a2206206994597C13D831ec7".lower()

@app.post("/api/hr/transactions/sync")
async def sync_etherscan_transactions():
    """등록된 지갑 주소에서 Etherscan API로 ERC-20 트랜잭션 동기화"""
    import httpx

    api_key = os.getenv("ETHERSCAN_API_KEY", "")
    if not api_key:
        raise HTTPException(500, "ETHERSCAN_API_KEY not configured")

    db = await get_db()

    # 등록된 지갑 조회
    w_rows = await db.execute("SELECT * FROM hr_wallets WHERE is_active=1")
    wallets = [dict(r) for r in await w_rows.fetchall()]
    if not wallets:
        await db.close()
        raise HTTPException(400, "등록된 지갑이 없습니다. 설정에서 지갑을 먼저 추가하세요.")

    # 기존 tx_hash 목록 (중복 방지)
    existing_rows = await db.execute("SELECT tx_hash FROM hr_transactions WHERE tx_hash IS NOT NULL AND tx_hash != ''")
    existing_hashes = set(r["tx_hash"] for r in await existing_rows.fetchall())

    added = 0
    errors = []

    async with httpx.AsyncClient(timeout=15) as client:
        for wallet in wallets:
            address = wallet["address"]
            url = (
                f"https://api.etherscan.io/v2/api"
                f"?chainid=1&module=account&action=tokentx"
                f"&address={address}"
                f"&startblock=0&endblock=99999999"
                f"&page=1&offset=1000"
                f"&sort=desc&apikey={api_key}"
            )
            try:
                resp = await client.get(url)
                data = resp.json()

                if data.get("status") != "1" or not data.get("result"):
                    continue

                # 허용할 토큰 contract (소문자)
                ALLOWED_CONTRACTS = {
                    USDT_CONTRACT: "USDT",                                          # Tether USDT
                    "0xa0b86991c6218b36c1d19d4a2e9eb0ce3606eb48": "USDC",         # USD Coin
                    "0x2be5e8c109e2197d077d13a82daead6a9b3433c5": "WTON",         # Wrapped TON
                }

                for tx in data["result"]:
                    tx_hash = tx["hash"]
                    if tx_hash in existing_hashes:
                        continue

                    # contract address로 필터 (스팸/가짜 토큰 제거)
                    contract = tx.get("contractAddress", "").lower()
                    if contract not in ALLOWED_CONTRACTS:
                        continue

                    token_symbol = ALLOWED_CONTRACTS[contract]
                    token_decimal = int(tx.get("tokenDecimal", 18))
                    amount = int(tx.get("value", 0)) / (10 ** token_decimal)

                    # 소액 필터 (1 미만 = 더스트/스팸)
                    if amount < 1:
                        continue

                    from_addr = tx["from"].lower()
                    to_addr = tx["to"].lower()
                    wallet_lower = address.lower()

                    # 방향 판별
                    direction = "in" if to_addr == wallet_lower else "out"
                    ts = datetime.utcfromtimestamp(int(tx["timeStamp"])).isoformat() + "Z"

                    note = f"{wallet['label']} {'입금' if direction == 'in' else '출금'}"

                    await db.execute(
                        "INSERT INTO hr_transactions (tx_hash, from_address, to_address, amount, token, status, timestamp, note) VALUES (?,?,?,?,?,?,?,?)",
                        (tx_hash, tx["from"], tx["to"], round(amount, 2), token_symbol, "confirmed", ts, note))
                    existing_hashes.add(tx_hash)
                    added += 1

            except Exception as e:
                errors.append(f"{wallet['label']}: {str(e)}")

    await db.commit()
    await db.close()

    msg = f"{added}건 동기화 완료"
    if errors:
        msg += f" (오류: {', '.join(errors)})"
    return {"added": added, "errors": errors, "message": msg}


@app.get("/api/hr/transactions/sync-status")
async def etherscan_sync_status():
    """동기화 가능 여부 확인"""
    api_key = os.getenv("ETHERSCAN_API_KEY", "")
    db = await get_db()
    w_rows = await db.execute("SELECT COUNT(*) as cnt FROM hr_wallets WHERE is_active=1")
    wallet_count = (await w_rows.fetchone())["cnt"]
    tx_rows = await db.execute("SELECT COUNT(*) as cnt FROM hr_transactions")
    tx_count = (await tx_rows.fetchone())["cnt"]
    await db.close()
    return {
        "api_key_set": bool(api_key),
        "wallet_count": wallet_count,
        "transaction_count": tx_count,
        "ready": bool(api_key) and wallet_count > 0,
    }


# ── Tax Simulation ──

@app.get("/api/hr/tax/simulate/{member_id}")
async def tax_simulate(member_id: int, year: int = 2026):
    db = await get_db()
    row = await db.execute("SELECT * FROM hr_members WHERE id=?", (member_id,))
    member = await row.fetchone()
    if not member:
        await db.close()
        raise HTTPException(404, "Member not found")

    p_rows = await db.execute("SELECT * FROM payrolls WHERE member_id=? AND year=? ORDER BY month", (member_id, year))
    payrolls = [dict(p) for p in await p_rows.fetchall()]
    total_payroll_krw = sum(p["krw_amount"] for p in payrolls)

    i_rows = await db.execute("SELECT * FROM incentives WHERE member_id=? AND year=?", (member_id, year))
    incentives = [dict(i) for i in await i_rows.fetchall()]
    total_incentive_krw = sum(i["krw_amount"] for i in incentives)

    annual_income = total_payroll_krw + total_incentive_krw
    tax_result = simulate_annual_tax(annual_income)
    monthly = monthly_tax_burden(annual_income)

    total_reserve_tok = sum(p["reserve_tokamak"] for p in payrolls)
    tokamak_price = 3200

    await db.close()
    return {
        "member": dict(member),
        "year": year,
        "payroll_income_krw": round(total_payroll_krw),
        "incentive_income_krw": round(total_incentive_krw),
        "annual_income_krw": round(annual_income),
        "tax": tax_result,
        "monthly_burden": monthly,
        "reserves": {"total_tokamak": round(total_reserve_tok, 4), "krw_value": round(total_reserve_tok * tokamak_price), "tokamak_price": tokamak_price},
        "payroll_details": payrolls,
    }


# ── Incentives ──

@app.get("/api/hr/incentives")
async def list_incentives(year: int = 2026, quarter: Optional[int] = None):
    db = await get_db()
    if quarter:
        rows = await db.execute("""
            SELECT i.*, m.name, m.role FROM incentives i
            JOIN hr_members m ON i.member_id = m.id
            WHERE i.year=? AND i.quarter=? ORDER BY m.name
        """, (year, quarter))
    else:
        rows = await db.execute("""
            SELECT i.*, m.name, m.role FROM incentives i
            JOIN hr_members m ON i.member_id = m.id
            WHERE i.year=? ORDER BY i.quarter, m.name
        """, (year,))
    result = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return result


# ── Transactions ──

@app.get("/api/hr/transactions")
async def list_hr_transactions(limit: int = 20):
    db = await get_db()
    rows = await db.execute("SELECT * FROM hr_transactions ORDER BY timestamp DESC LIMIT ?", (limit,))
    result = [dict(r) for r in await rows.fetchall()]
    await db.close()
    return result

class TransactionCreate(BaseModel):
    tx_hash: str = ""
    from_address: str = ""
    to_address: str = ""
    amount: float
    token: str = "USDT"
    status: str = "confirmed"
    timestamp: str = ""
    note: str = ""

@app.post("/api/hr/transactions")
async def create_hr_transaction(data: TransactionCreate):
    db = await get_db()
    ts = data.timestamp or datetime.now().isoformat()
    cursor = await db.execute(
        "INSERT INTO hr_transactions (tx_hash, from_address, to_address, amount, token, status, timestamp, note) VALUES (?,?,?,?,?,?,?,?)",
        (data.tx_hash, data.from_address, data.to_address, data.amount, data.token, data.status, ts, data.note))
    await db.commit()
    tid = cursor.lastrowid
    await db.close()
    return {"id": tid, "message": "Created"}

@app.delete("/api/hr/transactions/{tx_id}")
async def delete_hr_transaction(tx_id: int):
    db = await get_db()
    await db.execute("DELETE FROM hr_transactions WHERE id=?", (tx_id,))
    await db.commit()
    await db.close()
    return {"message": "Deleted"}


# ── Bulk Payroll History Upload (연도/월 포함 일괄 업로드) ──

@app.post("/api/hr/payroll/upload")
async def upload_payroll_history(file: UploadFile = File(...)):
    """
    급여이력 일괄 업로드.
    엑셀 컬럼: 연도, 월, 이름, USDT, 환율(선택), 세금(선택)
    환율/세금 없으면 0으로 저장, 상태는 paid.
    """
    import openpyxl, io

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    db = await get_db()
    added, updated, skipped = 0, 0, 0

    import re

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[2]:
            continue
        year = int(row[0])
        # "1월", "01", 1 등 다양한 형식 처리
        month_raw = str(row[1]).strip()
        month_digits = re.sub(r"[^0-9]", "", month_raw)
        if not month_digits:
            continue
        month = int(month_digits)
        name = str(row[2]).strip()
        usdt = float(row[3] or 0)
        rate = float(row[4]) if len(row) > 4 and row[4] else 0
        tax = float(row[5]) if len(row) > 5 and row[5] else 0

        krw = round(usdt * rate) if rate else 0
        net = krw - tax if krw else 0

        # 이름으로 멤버 매칭
        r = await db.execute("SELECT id FROM hr_members WHERE name=?", (name,))
        member = await r.fetchone()
        if not member:
            skipped += 1
            continue
        mid = member["id"]

        existing = await db.execute("SELECT id FROM payrolls WHERE member_id=? AND year=? AND month=?", (mid, year, month))
        ex = await existing.fetchone()

        if ex:
            await db.execute(
                "UPDATE payrolls SET usdt_amount=?, krw_rate=?, krw_amount=?, tax_simulated=?, net_pay_krw=?, status='paid' WHERE id=?",
                (usdt, rate, krw, tax, net, ex["id"]))
            updated += 1
        else:
            await db.execute(
                "INSERT INTO payrolls (member_id, year, month, usdt_amount, krw_rate, krw_amount, tax_simulated, reserve_tokamak, net_pay_krw, status) VALUES (?,?,?,?,?,?,?,0,?,?)",
                (mid, year, month, usdt, rate, krw, tax, net, "paid"))
            added += 1

    await db.commit()
    await db.close()
    return {"added": added, "updated": updated, "skipped": skipped, "message": f"{added}건 추가, {updated}건 업데이트, {skipped}건 스킵"}


@app.get("/api/hr/payroll/upload-template")
async def download_payroll_upload_template():
    """급여이력 일괄 업로드용 템플릿 (전 팀원 x 연월)"""
    import openpyxl, io
    from fastapi.responses import StreamingResponse

    db = await get_db()
    rows = await db.execute("SELECT name FROM hr_members ORDER BY is_active DESC, id")
    members = [r["name"] for r in await rows.fetchall()]
    await db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "급여이력"
    ws.append(["연도", "월", "이름", "USDT", "환율", "세금(KRW)"])

    # 샘플: 현재 연월 한 줄씩
    from datetime import datetime as dt
    now = dt.now()
    for m in members:
        ws.append([now.year, now.month, m, 0, 0, 0])

    for col, w in [("A",8),("B",6),("C",15),("D",12),("E",10),("F",12)]:
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=payroll_history_template.xlsx"})


# ── Bulk Payroll Calculation (엑셀 업로드) ──

@app.post("/api/hr/calculate/preview")
async def calculate_payroll_preview(file: UploadFile = File(...)):
    """
    엑셀 업로드 → 미리보기 (저장 안 함).
    엑셀 컬럼: 이름, USDT, 환율, 부양가족수(기본1), 8-20세자녀수(기본0)
    """
    import openpyxl, io
    from tax_calculator import calculate_tax

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    results = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        usdt = float(row[1] or 0)
        rate = float(row[2] or 0)
        dependents = int(row[3]) if len(row) > 3 and row[3] else 1
        children = int(row[4]) if len(row) > 4 and row[4] else 0

        krw = round(usdt * rate)
        tax_result = calculate_tax(krw, dependents, children)
        tax = tax_result["total_tax_100"]
        net = krw - tax

        results.append({
            "name": name,
            "usdt_amount": usdt,
            "krw_rate": rate,
            "krw_amount": krw,
            "income_tax": tax_result["income_tax_100"],
            "local_tax": tax_result["local_tax_100"],
            "tax_total": tax,
            "net_pay_krw": net,
            "dependents": dependents,
            "children": children,
        })

    return {"results": results}


@app.post("/api/hr/calculate/save")
async def calculate_payroll_save(data: dict):
    """
    미리보기 결과를 payrolls 테이블에 저장.
    data: { year, month, status, results: [...preview results] }
    """
    db = await get_db()
    year = data["year"]
    month = data["month"]
    status = data.get("status", "estimated")
    saved = 0

    for r in data["results"]:
        # 이름으로 멤버 매칭
        row = await db.execute("SELECT id FROM hr_members WHERE name=? AND is_active=1", (r["name"],))
        member = await row.fetchone()
        if not member:
            continue
        mid = member["id"]

        # 기존 데이터 확인
        existing = await db.execute("SELECT id FROM payrolls WHERE member_id=? AND year=? AND month=?", (mid, year, month))
        ex = await existing.fetchone()

        if ex:
            await db.execute(
                "UPDATE payrolls SET usdt_amount=?, krw_rate=?, krw_amount=?, tax_simulated=?, net_pay_krw=?, status=? WHERE id=?",
                (r["usdt_amount"], r["krw_rate"], r["krw_amount"], r["tax_total"], r["net_pay_krw"], status, ex["id"]))
        else:
            await db.execute(
                "INSERT INTO payrolls (member_id, year, month, usdt_amount, krw_rate, krw_amount, tax_simulated, reserve_tokamak, net_pay_krw, status) VALUES (?,?,?,?,?,?,?,0,?,?)",
                (mid, year, month, r["usdt_amount"], r["krw_rate"], r["krw_amount"], r["tax_total"], r["net_pay_krw"], status))
        saved += 1

    await db.commit()
    await db.close()
    return {"saved": saved, "message": f"{saved}명 급여 데이터 저장 완료"}


@app.get("/api/hr/calculate/template")
async def download_payroll_template():
    """급여계산용 엑셀 템플릿 다운로드"""
    import openpyxl, io
    from fastapi.responses import StreamingResponse

    db = await get_db()
    rows = await db.execute("SELECT name, monthly_usdt FROM hr_members WHERE is_active=1 ORDER BY id")
    members = [dict(r) for r in await rows.fetchall()]
    await db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "급여계산"
    ws.append(["이름", "USDT", "환율", "부양가족수", "8-20세자녀수"])

    for m in members:
        ws.append([m["name"], m["monthly_usdt"], 0, 1, 0])

    # 컬럼 너비
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=payroll_template.xlsx"})


# ── Tax Calculator (간이세액표 기반) ──

class TaxCalcRequest(BaseModel):
    monthly_salary_krw: int
    num_dependents: int = 1
    num_children_8_20: int = 0

@app.post("/api/hr/tax/calculate")
async def tax_calculate(data: TaxCalcRequest):
    """2026 간이세액표 기반 소득세 + 지방소득세 계산"""
    result = calculate_tax(data.monthly_salary_krw, data.num_dependents, data.num_children_8_20)
    return result


# ── 한국은행 ECOS 환율 API ──

@app.get("/api/hr/exchange-rate")
async def get_exchange_rate(date: str = ""):
    """
    한국은행 ECOS API에서 원/달러 종가(15:30) 조회.
    date: YYYYMMDD 또는 YYYY-MM-DD 형식. 미입력 시 최근 영업일.
    """
    import httpx

    ecos_key = os.getenv("ECOS_API_KEY", "")
    if not ecos_key:
        raise HTTPException(500, "ECOS_API_KEY not configured")

    # 날짜 정리
    if date:
        clean_date = date.replace("-", "")
    else:
        # 최근 7일 범위로 조회 (주말/공휴일 대비)
        from datetime import timedelta
        today = datetime.now()
        clean_date = (today - timedelta(days=7)).strftime("%Y%m%d")
        end_date = today.strftime("%Y%m%d")

    if date:
        start = clean_date
        end = clean_date
    else:
        start = clean_date
        end = end_date

    # 731Y003 / D / 0000003 = 원/달러(종가 15:30)
    url = f"https://ecos.bok.or.kr/api/StatisticSearch/{ecos_key}/JSON/kr/1/10/731Y003/D/{start}/{end}/0000003"

    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(url)
            data = resp.json()
    except Exception as e:
        raise HTTPException(502, f"ECOS API error: {e}")

    if "StatisticSearch" not in data:
        error_msg = data.get("RESULT", {}).get("MESSAGE", "Unknown error")
        raise HTTPException(404, f"환율 데이터 없음: {error_msg}")

    rows = data["StatisticSearch"]["row"]

    if date:
        # 특정 날짜 요청 — 해당 날짜 or 가장 가까운 이전 영업일
        result_row = rows[-1] if rows else None
    else:
        # 최근 영업일
        result_row = rows[-1] if rows else None

    if not result_row:
        raise HTTPException(404, "해당 날짜의 환율 데이터가 없습니다")

    rate = float(result_row["DATA_VALUE"])
    rate_date = result_row["TIME"]  # YYYYMMDD

    return {
        "rate": rate,
        "date": f"{rate_date[:4]}-{rate_date[4:6]}-{rate_date[6:8]}",
        "source": "한국은행 ECOS",
        "item": "원/달러(종가 15:30)",
    }


@app.get("/api/hr/exchange-rate/prev-day")
async def get_exchange_rate_prev_day(date: str):
    """
    지급일 전날 종가 환율 조회.
    date: YYYY-MM-DD 형식의 지급일. 전날부터 최대 7일 이전까지 탐색하여 가장 가까운 영업일 종가 반환.
    """
    import httpx
    from datetime import timedelta

    ecos_key = os.getenv("ECOS_API_KEY", "")
    if not ecos_key:
        raise HTTPException(500, "ECOS_API_KEY not configured")

    pay_date = datetime.strptime(date.replace("-", ""), "%Y%m%d")
    end = pay_date - timedelta(days=1)
    start = end - timedelta(days=7)

    url = f"https://ecos.bok.or.kr/api/StatisticSearch/{ecos_key}/JSON/kr/1/10/731Y003/D/{start.strftime('%Y%m%d')}/{end.strftime('%Y%m%d')}/0000003"

    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(url)
            data = resp.json()
    except Exception as e:
        raise HTTPException(502, f"ECOS API error: {e}")

    if "StatisticSearch" not in data:
        raise HTTPException(404, "환율 데이터 없음")

    rows = data["StatisticSearch"]["row"]
    if not rows:
        raise HTTPException(404, "해당 기간의 환율 데이터가 없습니다")

    last = rows[-1]
    rate_date = last["TIME"]
    return {
        "rate": float(last["DATA_VALUE"]),
        "date": f"{rate_date[:4]}-{rate_date[4:6]}-{rate_date[6:8]}",
        "source": "한국은행 ECOS (전일 종가)",
    }

@app.get("/api/hr/exchange-rate/range")
async def get_exchange_rate_range(start: str, end: str):
    """날짜 범위의 환율 조회. start/end: YYYYMMDD 또는 YYYY-MM-DD"""
    import httpx

    ecos_key = os.getenv("ECOS_API_KEY", "")
    if not ecos_key:
        raise HTTPException(500, "ECOS_API_KEY not configured")

    s = start.replace("-", "")
    e = end.replace("-", "")
    url = f"https://ecos.bok.or.kr/api/StatisticSearch/{ecos_key}/JSON/kr/1/100/731Y003/D/{s}/{e}/0000003"

    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get(url)
            data = resp.json()
    except Exception as ex:
        raise HTTPException(502, f"ECOS API error: {ex}")

    if "StatisticSearch" not in data:
        return {"rates": []}

    rows = data["StatisticSearch"]["row"]
    return {
        "rates": [
            {
                "date": f"{r['TIME'][:4]}-{r['TIME'][4:6]}-{r['TIME'][6:8]}",
                "rate": float(r["DATA_VALUE"]),
            }
            for r in rows
        ],
        "source": "한국은행 ECOS",
    }


# ── Payslip PDF Generator ──

class PayslipRequest(BaseModel):
    contractor_name: str
    erc20_address: str = ""
    transaction_url: str = ""
    payment_year: int
    payment_month: int
    service_fee_usdt: float
    exchange_rate: float
    income_tax_krw: int
    local_tax_krw: int
    total_tax_krw: int
    tax_percentage: int = 100

@app.get("/api/hr/generate-payslip")
async def generate_payslip_from_payroll(member_id: int, year: int, month: int):
    """급여 데이터 기반 Payslip PDF 자동 생성"""
    from fastapi.responses import Response
    from payslip_pdf import generate_payslip_pdf
    from tax_calculator import calculate_tax

    db = await get_db()
    # 멤버 정보
    m_row = await db.execute("SELECT * FROM hr_members WHERE id=?", (member_id,))
    member = await m_row.fetchone()
    if not member:
        await db.close()
        raise HTTPException(404, "Member not found")

    # 급여 데이터
    p_row = await db.execute("SELECT * FROM payrolls WHERE member_id=? AND year=? AND month=?", (member_id, year, month))
    payroll = await p_row.fetchone()
    if not payroll:
        await db.close()
        raise HTTPException(404, "Payroll not found")

    # 지갑 주소 (첫 번째)
    w_row = await db.execute("SELECT address FROM member_wallets WHERE member_id=? LIMIT 1", (member_id,))
    wallet = await w_row.fetchone()
    erc20_address = wallet["address"] if wallet else member["wallet_address"] or ""

    await db.close()

    p = dict(payroll)
    rate = p["krw_rate"] or 0
    krw = p["krw_amount"] or 0

    # 세금 분리 (소득세/지방소득세)
    if rate > 0 and krw > 0:
        tax_detail = calculate_tax(int(krw))
        income_tax = tax_detail["income_tax_100"]
        local_tax = tax_detail["local_tax_100"]
        total_tax = tax_detail["total_tax_100"]
    else:
        income_tax = 0
        local_tax = 0
        total_tax = int(p["tax_simulated"] or 0)

    pdf_bytes = generate_payslip_pdf(
        contractor_name=member["name"],
        erc20_address=erc20_address,
        transaction_url="",
        payment_year=year,
        payment_month=month,
        service_fee_usdt=p["usdt_amount"],
        exchange_rate=rate,
        income_tax_krw=income_tax,
        local_tax_krw=local_tax,
        total_tax_krw=total_tax,
        tax_percentage=100,
    )

    filename = f"payslip_{member['name']}_{year}{str(month).zfill(2)}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@app.post("/api/hr/payslip/pdf")
async def generate_payslip(data: PayslipRequest):
    from fastapi.responses import Response
    from payslip_pdf import generate_payslip_pdf

    pdf_bytes = generate_payslip_pdf(
        contractor_name=data.contractor_name,
        erc20_address=data.erc20_address,
        transaction_url=data.transaction_url,
        payment_year=data.payment_year,
        payment_month=data.payment_month,
        service_fee_usdt=data.service_fee_usdt,
        exchange_rate=data.exchange_rate,
        income_tax_krw=data.income_tax_krw,
        local_tax_krw=data.local_tax_krw,
        total_tax_krw=data.total_tax_krw,
        tax_percentage=data.tax_percentage,
    )

    filename = f"payslip_{data.contractor_name}_{data.payment_year}{str(data.payment_month).zfill(2)}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/hr/market/tokamak")
async def tokamak_price():
    return {"token": "TOKAMAK", "price_krw": 3200, "source": "mock", "timestamp": datetime.now().isoformat()}

@app.get("/api/hr/market/usdt")
async def usdt_rate():
    return {"pair": "USDT/KRW", "rate": 1352.50, "source": "mock", "timestamp": datetime.now().isoformat()}


if __name__ == "__main__":
    import uvicorn
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=8001)
    args = parser.parse_args()
    uvicorn.run(app, host="0.0.0.0", port=args.port)
